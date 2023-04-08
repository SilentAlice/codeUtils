#!/usr/local/bin/python3

## Using this to combine all xlsx data together

# system modules
import datetime
import openpyxl
import sys
import re
from os import chdir
from pathlib import Path

WORKDIR="/home/alice/Win/Temp"
DIRNAME="Oertime_batch_ExAid"

def __debug_output_envs():
    print("WORKDIR:" + WORKDIR)
    print("DIRNAME:" + DIRNAME)
    print("======End======")

def __process_date():
    # Using current date as prefix
    today = datetime.date.today()
    prefix = "{:4d}{:0>2d}{:0>2d}".format(today.year, today.month, today.day)
    return prefix

def __is_filename_template(filename: str):
    return re.match(".*template.xlsx", filename, re.I)

def __process_single_file(singleName: str):
    from openpyxl.chart import Reference
    wbSingle = openpyxl.load_workbook(singleName)
    st = wbSingle.active
    rows = list(st.rows)

    # row 1 and 2 are titles
    for r in list(st.rows)[2:]:
        # col 1 is reserved
        if (r[1].value is None):
            # end of data, return row[2:]
            wbSingle.close()
            return rows[2:r[1].row - 1]
    return rows

def __format_and_check_online(row):
    name = str(row[2].value)
    # check none cells
    for idx, v in enumerate(row[1:9]):
        if v.value is None:
            print(name + " has none data: col->" + chr(ord('B') + idx))
            return None

    sD = row[3].value
    eD = row[5].value

    # startDay must be same with endDay
    if sD != eD:
        print(name +  " startDay:" +  sD.strftime("%Y-%m-%d") +
              " != endDay:" + eD.strftime("%Y-%m-%d"))
        return None

    # year must be this year
    if sD.year != datetime.date.today().year:
        print(name + "not this year " + str(sD))
        return None

    # get needed times
    times = [row[4]] + list(row[6:9])
    # convert chinese :
    times = map(lambda t: str(t.value).replace('：', ':'), times)
    # convert to iso time
    times = map(lambda t: t if len(t) == 5 else '0' + t, times)
    times = list(map(lambda t: datetime.time.fromisoformat(t), times))
    [sT, eT, sR, eR] = times

    # not later than 10:30
    askT = datetime.time.fromisoformat("10:30")
    if sT > askT:
        print(name + " startTime:  " + str(sT.strftime("%H:%M") + " Wrong"))
        return None

    # no larger than 8 hours one day
    diff = eT.minute - sT.minute - (eR.minute - sR.minute)
    diff = eT.hour - sT.hour - (eR.hour - sR.housr)

    if diff != 8:
        print(name + " overtime is not 8 hours:" + str(diff) + " ->", end="")
        [print(" " + t.strftime("%H:%M"), end="") for t in times]
        print("")
        return None

    times = [t.strftime("%H:%M") for t in times]

    row = [d.value for d in row[0:12]]
    row = row[0:3] + [sD] + [times[0]] + [eD] + times[1:4] + row[9:12]

    return row

def __format_and_check_lines(rows):
    if len(rows) < 1:
        return None

    setData = set()
    formatedRows = list()

    for r in rows:
        outData = __format_and_check_online(r)
        if outData is None:
            # just leave out this record
            continue

        # check if duplicated record and leave out
        lenBefore = len(setData)
        # using startday as unique key
        setData.add(outData[3].strftime("%d"))
        if lenBefore == len(setData):
            print(outData[2] + " duplicated records")
            continue

        formatedRows += [outData]

    return formatedRows

def __update_final_sheet(final, rows):
    # get last non-null row
    for idx, r in enumerate(final.rows):
        if r[1].value is None:
            newIdx = idx
            break

    for idRow, r in enumerate(rows):
        # set start day and end day to mm-dd-yy format
        final.cell(row = idRow + idx + 1, column = 4).number_format = "mm-dd-yy"
        final.cell(row = idRow + idx + 1, column = 6).number_format = "mm-dd-yy"

        for idCol, c in enumerate(r):
            # write new value to final sheet, row and col start from 1
            final.cell(row = idRow + idx + 1, column = idCol + 1).value = c

def __addto_final_file(final, rows):
    rows = __format_and_check_lines(rows)
    __update_final_sheet(final, rows)

def __process_overtime():
    path = Path('.')
    # using YYYYMMDD as prefix
    prefix = __process_date()

    for f in path.rglob("*.*"):
        if __is_filename_template(f.name):
            workbook = openpyxl.load_workbook(f.name)
            final = workbook.active
            break

    for f in path.rglob("*.*"):
        if not __is_filename_template(f.name):
            data = __process_single_file(f.absolute())
            __addto_final_file(final, data)

    workbook.save(prefix + DIRNAME + ".xlsx")
    workbook.close()

def main():
    __debug_output_envs()

    # change to workdir (where xlsx is)
    chdir(WORKDIR + DIRNAME)
    __process_overtime()

main()
